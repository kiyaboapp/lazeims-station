"""Excel export for station marks and attendance data.

Uses openpyxl to produce .xlsx workbooks. Each function accepts an open
SQLite connection and an output file path.
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def _write_header(ws, headers: list[str]) -> None:
    """Write a styled header row at row 1."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def export_marks_excel(conn: sqlite3.Connection, output_path: str | Path) -> None:
    """Export all marks (total_marks + item_marks) to an Excel workbook.

    Creates one sheet per subject. Columns include Student ID, full name,
    Paper Type, Total Marks Obtained, plus individual item marks columns
    when item-level data exists.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Get all subjects that have marks
    subjects = conn.execute(
        "SELECT DISTINCT s.subject_code, sub.name "
        "FROM total_marks tm "
        "JOIN students s ON s.student_id = tm.student_id "
        "JOIN student_subjects ss ON ss.student_id = s.student_id AND ss.subject_code = tm.subject_code "
        "JOIN subjects sub ON sub.subject_code = tm.subject_code "
        "ORDER BY tm.subject_code"
    ).fetchall()

    if not subjects:
        # Create an empty sheet if no data
        ws = wb.create_sheet("No Data")
        ws.cell(row=1, column=1, value="No marks data available")
        wb.save(str(output_path))
        return

    for subj in subjects:
        sc = subj["subject_code"]
        subj_name = subj["name"] or sc
        # Sheet name max 31 chars in Excel
        sheet_title = f"{sc} - {subj_name}"[:31]
        ws = wb.create_sheet(title=sheet_title)

        # Get questions for this subject (for item marks columns)
        questions = conn.execute(
            "SELECT id, paper_type, question_number, max_marks "
            "FROM questions WHERE subject_code = ? ORDER BY paper_type, question_number",
            (sc,)
        ).fetchall()

        # Base headers
        headers = ["Student ID", "Full Name", "Paper Type", "Total Marks Obtained"]

        # Add question columns per paper type if item marks exist
        q_map: dict[str, list] = {}  # paper_type -> [(question_number, question_id)]
        for q in questions:
            pt = q["paper_type"]
            if pt not in q_map:
                q_map[pt] = []
            q_map[pt].append((q["question_number"], q["id"]))

        # We will add item marks as additional columns only if they exist
        has_items = conn.execute(
            "SELECT 1 FROM item_marks im "
            "JOIN questions q ON q.id = im.question_id "
            "WHERE q.subject_code = ? LIMIT 1",
            (sc,)
        ).fetchone() is not None

        if has_items:
            # Collect all distinct question numbers across paper types
            all_q_numbers = sorted(set(
                q["question_number"] for q in questions
            ))
            for qn in all_q_numbers:
                headers.append(f"Q{qn}")

        _write_header(ws, headers)

        # Fetch marks data
        marks_rows = conn.execute(
            "SELECT tm.student_id, s.first_name, s.middle_name, s.surname, "
            "tm.paper_type, tm.total_marks_obtained "
            "FROM total_marks tm "
            "JOIN students s ON s.student_id = tm.student_id "
            "WHERE tm.subject_code = ? "
            "ORDER BY tm.paper_type, tm.student_id",
            (sc,)
        ).fetchall()

        row_idx = 2
        for mark in marks_rows:
            full_name = (
                (mark["first_name"] or "").strip().upper()
                + (" " + (mark["middle_name"] or "").strip().upper() if mark["middle_name"] else "")
                + " " + (mark["surname"] or "").strip().upper()
            ).strip()

            ws.cell(row=row_idx, column=1, value=mark["student_id"])
            ws.cell(row=row_idx, column=2, value=full_name)
            ws.cell(row=row_idx, column=3, value=mark["paper_type"])
            ws.cell(row=row_idx, column=4, value=float(mark["total_marks_obtained"]))

            # Item marks
            if has_items:
                all_q_numbers = sorted(set(
                    q["question_number"] for q in questions
                ))
                # Get item marks for this student and paper
                item_rows = conn.execute(
                    "SELECT q.question_number, im.marks_obtained "
                    "FROM item_marks im "
                    "JOIN questions q ON q.id = im.question_id "
                    "WHERE im.student_id = ? AND q.subject_code = ? AND q.paper_type = ?",
                    (mark["student_id"], sc, mark["paper_type"])
                ).fetchall()
                item_dict = {r["question_number"]: float(r["marks_obtained"]) for r in item_rows}

                for col_offset, qn in enumerate(all_q_numbers):
                    val = item_dict.get(qn)
                    if val is not None:
                        ws.cell(row=row_idx, column=5 + col_offset, value=val)

            row_idx += 1

    wb.save(str(output_path))


def export_attendance_excel(conn: sqlite3.Connection, output_path: str | Path) -> None:
    """Export all attendance records to an Excel workbook.

    Single sheet with columns: Student ID, Full Name, Centre Number,
    Subject Code, Paper Type, Is Present, Source.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["Student ID", "Full Name", "Centre Number", "Subject Code",
               "Paper Type", "Is Present", "Source"]
    _write_header(ws, headers)

    rows = conn.execute(
        "SELECT a.student_id, s.first_name, s.middle_name, s.surname, "
        "s.centre_number, a.subject_code, a.paper_type, a.is_present, a.source "
        "FROM attendance a "
        "JOIN students s ON s.student_id = a.student_id "
        "ORDER BY s.centre_number, a.subject_code, a.paper_type, a.student_id"
    ).fetchall()

    if not rows:
        ws.cell(row=2, column=1, value="No attendance data available")
        wb.save(str(output_path))
        return

    for row_idx, row in enumerate(rows, start=2):
        full_name = (
            (row["first_name"] or "").strip().upper()
            + (" " + (row["middle_name"] or "").strip().upper() if row["middle_name"] else "")
            + " " + (row["surname"] or "").strip().upper()
        ).strip()

        ws.cell(row=row_idx, column=1, value=row["student_id"])
        ws.cell(row=row_idx, column=2, value=full_name)
        ws.cell(row=row_idx, column=3, value=row["centre_number"])
        ws.cell(row=row_idx, column=4, value=row["subject_code"])
        ws.cell(row=row_idx, column=5, value=row["paper_type"])
        ws.cell(row=row_idx, column=6, value="YES" if row["is_present"] else "NO")
        ws.cell(row=row_idx, column=7, value=row["source"] or "")

    wb.save(str(output_path))


def export_all_data_excel(conn: sqlite3.Connection, output_path: str | Path) -> None:
    """Combined export: attendance sheet + per-subject marks sheets.

    First sheet is attendance, followed by one marks sheet per subject.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ---- Attendance sheet ----
    ws_att = wb.create_sheet(title="Attendance")
    att_headers = ["Student ID", "Full Name", "Centre Number", "Subject Code",
                   "Paper Type", "Is Present", "Source"]
    _write_header(ws_att, att_headers)

    att_rows = conn.execute(
        "SELECT a.student_id, s.first_name, s.middle_name, s.surname, "
        "s.centre_number, a.subject_code, a.paper_type, a.is_present, a.source "
        "FROM attendance a "
        "JOIN students s ON s.student_id = a.student_id "
        "ORDER BY s.centre_number, a.subject_code, a.paper_type, a.student_id"
    ).fetchall()

    for row_idx, row in enumerate(att_rows, start=2):
        full_name = (
            (row["first_name"] or "").strip().upper()
            + (" " + (row["middle_name"] or "").strip().upper() if row["middle_name"] else "")
            + " " + (row["surname"] or "").strip().upper()
        ).strip()
        ws_att.cell(row=row_idx, column=1, value=row["student_id"])
        ws_att.cell(row=row_idx, column=2, value=full_name)
        ws_att.cell(row=row_idx, column=3, value=row["centre_number"])
        ws_att.cell(row=row_idx, column=4, value=row["subject_code"])
        ws_att.cell(row=row_idx, column=5, value=row["paper_type"])
        ws_att.cell(row=row_idx, column=6, value="YES" if row["is_present"] else "NO")
        ws_att.cell(row=row_idx, column=7, value=row["source"] or "")

    if not att_rows:
        ws_att.cell(row=2, column=1, value="No attendance data")

    # ---- Marks sheets (per subject) ----
    subjects = conn.execute(
        "SELECT DISTINCT tm.subject_code, sub.name "
        "FROM total_marks tm "
        "JOIN subjects sub ON sub.subject_code = tm.subject_code "
        "ORDER BY tm.subject_code"
    ).fetchall()

    for subj in subjects:
        sc = subj["subject_code"]
        subj_name = subj["name"] or sc
        sheet_title = f"Marks - {sc}"[:31]
        ws = wb.create_sheet(title=sheet_title)

        # Get questions for this subject
        questions = conn.execute(
            "SELECT id, paper_type, question_number, max_marks "
            "FROM questions WHERE subject_code = ? ORDER BY paper_type, question_number",
            (sc,)
        ).fetchall()

        headers = ["Student ID", "Full Name", "Paper Type", "Total Marks"]

        has_items = conn.execute(
            "SELECT 1 FROM item_marks im "
            "JOIN questions q ON q.id = im.question_id "
            "WHERE q.subject_code = ? LIMIT 1",
            (sc,)
        ).fetchone() is not None

        all_q_numbers: list[str] = []
        if has_items:
            all_q_numbers = sorted(set(
                q["question_number"] for q in questions
            ))
            for qn in all_q_numbers:
                headers.append(f"Q{qn}")

        _write_header(ws, headers)

        marks_rows = conn.execute(
            "SELECT tm.student_id, s.first_name, s.middle_name, s.surname, "
            "tm.paper_type, tm.total_marks_obtained "
            "FROM total_marks tm "
            "JOIN students s ON s.student_id = tm.student_id "
            "WHERE tm.subject_code = ? "
            "ORDER BY tm.paper_type, tm.student_id",
            (sc,)
        ).fetchall()

        for row_idx, mark in enumerate(marks_rows, start=2):
            full_name = (
                (mark["first_name"] or "").strip().upper()
                + (" " + (mark["middle_name"] or "").strip().upper() if mark["middle_name"] else "")
                + " " + (mark["surname"] or "").strip().upper()
            ).strip()

            ws.cell(row=row_idx, column=1, value=mark["student_id"])
            ws.cell(row=row_idx, column=2, value=full_name)
            ws.cell(row=row_idx, column=3, value=mark["paper_type"])
            ws.cell(row=row_idx, column=4, value=float(mark["total_marks_obtained"]))

            if has_items:
                item_rows = conn.execute(
                    "SELECT q.question_number, im.marks_obtained "
                    "FROM item_marks im "
                    "JOIN questions q ON q.id = im.question_id "
                    "WHERE im.student_id = ? AND q.subject_code = ? AND q.paper_type = ?",
                    (mark["student_id"], sc, mark["paper_type"])
                ).fetchall()
                item_dict = {r["question_number"]: float(r["marks_obtained"]) for r in item_rows}
                for col_offset, qn in enumerate(all_q_numbers):
                    val = item_dict.get(qn)
                    if val is not None:
                        ws.cell(row=row_idx, column=5 + col_offset, value=val)

    if not subjects:
        ws_marks = wb.create_sheet(title="Marks")
        ws_marks.cell(row=1, column=1, value="No marks data available")

    wb.save(str(output_path))
